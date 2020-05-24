import requests, zipfile, io,os,datetime,csv
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import math
import pandas
import shutil 
import os 
from openpyxl import load_workbook
import xlsxwriter
from bs4 import BeautifulSoup
import csv
import time

if not os.path.exists(os.getcwd()+"//company"):
    os.mkdir("company")
if not os.path.exists(os.getcwd()+"//temp"):
    os.mkdir("temp")
if not os.path.exists(os.getcwd()+"//turnover"):
    os.mkdir("turnover")
if not os.path.exists(os.getcwd()+"//nifty"):
    os.mkdir("nifty")



# for Driver
local_dir = os.getcwd()+'\geckodriver.exe'
fp = webdriver.FirefoxProfile()
fp.set_preference("browser.download.folderList",2)
fp.set_preference("browser.download.manager.showWhenStarting",False)
fp.set_preference("browser.download.dir",os.getcwd()+"\\temp")
fp.set_preference("browser.helperApps.neverAsk.saveToDisk",r"text/csv (officially registered type), application/csv, text/x-csv, application/x-csv, text/x-comma-separated-values, text/comma-separated-values")
driver = webdriver.Firefox(executable_path =local_dir,firefox_profile=fp)
# driver2 = webdriver.Firefox(executable_path =local_dir,firefox_profile=fp)

# recored to update
rcd_update = 10

nifyty_header = ["DATE","OPEN ","HIGH","LOW","CLOSE","% CHG","TURNOVER(Cr)","% CHG","5 DAY AVG TURNOVER","VOLUME","CHG IN VOL","% CHG","5 DAY AVG VOL","","OI","CHG_IN_OI","% CHG","5 DAY AVG OI","VOL(CONTRACTS)","CHG IN VOL","% CHG","5 DAY AVG VOL","","OPT OI","COI","% CHG","5 DAY AVG OI","VOL(CONTRACTS)","CHG IN VOL","% CHG","5 DAY AVG VOL"]
company_name =["INFRATEL","BRITANNIA","HCLTECH","ITC","SHREECEM","EICHERMOT","CIPLA","SUNPHARMA","COALINDIA","BHARTIARTL","HINDUNILVR","TITAN","ONGC","LT","TECHM","GAIL","ASIANPAINT","DRREDDY","NTPC","WIPRO","POWERGRID","NESTLEIND","BAJAJ-AUTO","TATAMOTORS","BPCL","ULTRACEMCO","IOC","VEDL","MARUTI","UPL","BAJAJFINSV","HEROMOTOCO","HINDALCO","ADANIPORTS","SBIN","TATASTEEL","RELIANCE","JSWSTEEL","INFY","M&M","GRASIM","TCS","HDFCBANK","KOTAKBANK","BAJFINANCE","AXISBANK","HDFC","ICICIBANK","ZEEL","INDUSINDBK"]
months = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]
feildName=["DATE","OPEN","HIGH","LOW","CLOSE","%CHG","VWAP","5 DAY VWAP","TURNOVER","5 DAY AVG TURNOVER","VOLUME","CHNG IN VOL","%CHANGE","5 DAY AVG VOL","","OL","CHG_IN_OI","%CHANGE","5 DAYS AVG OI","TURNOVER","5 DAY AVG TURNOVER","VOL(CONTRACTS)","CHG IN VOL","% CHG","5 DAYS AVG VOL","","OPEN_INT","CHG_IN_OI","% CHANGE","5 DAYS AVG OI","TURNOVER","5 DAY AVG TURNOVER","VOL(CONTRACTS)","CHG IN VOL","%CHANGE","5 DAYS AVG VOL"]

# for retry of request of download
retry_strategy = Retry(
    total=3,
    status_forcelist=[429, 500, 502, 503, 504],
    method_whitelist=["HEAD", "GET", "OPTIONS"]
)
adapter = HTTPAdapter(max_retries=retry_strategy)
http = requests.Session()
http.mount("https://", adapter)
http.mount("http://", adapter)

# company_name = ["SBIN"]
feildName=["DATE","OPEN","HIGH","LOW","CLOSE","%CHG","VWAP","5 DAY VWAP","TURNOVER","5 DAY AVG TURNOVER","VOLUME","CHNG IN VOL","%CHANGE","5 DAY AVG VOL","","OL","CHG_IN_OI","%CHANGE","5 DAYS AVG OI","TURNOVER","5 DAY AVG TURNOVER","VOL(CONTRACTS)","CHG IN VOL","% CHG","5 DAYS AVG VOL","","OPEN_INT","CHG_IN_OI","% CHANGE","5 DAYS AVG OI","TURNOVER","5 DAY AVG TURNOVER","VOL(CONTRACTS)","CHG IN VOL","%CHANGE","5 DAYS AVG VOL"]


datetoday = datetime.datetime.now()
day_shift =-1
total_found = 0
fileIndex =[]
daylist = []
while(total_found!=rcd_update+4):

#     DateMaker and FileNameMaker
    day_shift +=1
    new_date = datetoday - datetime.timedelta(days=day_shift)
    new_day = (new_date).day
    if(new_day<10):
        new_day = "0"+ str(new_day)
    new_month = months[(new_date).month-1]
    new_year = (new_date).year
    filename = "fo"+str(new_day)+new_month+str(new_year)+"bhav.csv"
    
    
    addr = "https://www1.nseindia.com/content/historical/DERIVATIVES/"+str(new_year)+r'/'+ new_month + r'/'+filename+'.zip'
    if(os.path.exists(os.getcwd()+"\\temp\\"+filename)):
        total_found +=1
        daylist.append(str(new_day)+"-"+str(new_month)+"-"+str(new_year))
        print(str(new_day)+"-"+str(new_month)+"-"+str(new_year), "day file exists")
        fileIndex.append(filename)
    else:
        p = 1
        while p:
            try:
                r = requests.get(addr,timeout=5)
                p =0
            except:
                print("failed to fetch slow conection retrying")
                pass
#             r = requests.get(addr,timeout=20)
        if(r.status_code == 200):
            z = zipfile.ZipFile(io.BytesIO(r.content))            
            z.extractall()
            total_found +=1
            fileIndex.append(filename)
            daylist.append(str(new_day)+"-"+str(new_month)+"-"+str(new_year))
            print(str(new_day)+"-"+str(new_month)+"-"+str(new_year), "day file Downloaded")
            # To move the files
            dest = shutil.move(os.getcwd()+"\\"+filename, os.getcwd()+"\\temp\\"+filename)
            
print("all bhav files downloaded")
# extra_day = str(daylist[-1])
# daylist = daylist[:-1]
# extra_day_file = fileIndex[-1] 
# fileIndex = fileIndex[:-1]

if (months.index(daylist[-1][3:6])+1)<9 :
    month_start =  "0"+str(months.index(daylist[-1][3:6])+1)
else:
    month_start = str(months.index(daylist[-1][3:6])+1)   
if (months.index(daylist[0][3:6]))<9 :
    month_end =  "0"+str(months.index(daylist[0][3:6])+1)
else:
    month_end = str(months.index(daylist[0][3:6])+1)   

    
end_day =   daylist[-1][0:3]+ month_start+ daylist[-1][6:]
start_day = daylist[0][0:3]+month_end+ daylist[0][6:]




all_days_df = []
for i,filename in enumerate(fileIndex):
        df = pandas.read_csv(os.getcwd()+"\\temp\\"+filename)
        all_days_df.append(df)


driver.get("https://www1.nseindia.com/products/content/equities/equities/eq_security.htm")
for company in company_name:
    
#     to decide whether to create or  append
    create_new =1
    continue_from = 0
#     append_ = 0 
    if (os.path.exists(os.getcwd()+"\\company\\"+company+".xlsx")):
        edit_date = daylist[rcd_update-1]
        append_ = 1
        create_new = 0
        ty = pandas.read_excel(os.getcwd()+"\\company\\"+company+".xlsx").iloc[:,0]
        for i in range(ty.size):
                if(i>2 and ty.iloc[i].upper() == edit_date):
                    append_ = 0
                    continue_from = i
                    break
        
# #     for turnover in cash of each company
#     driver2.get(r"https://www1.nseindia.com/companytracker/cmtracker.jsp?symbol="+company+"&cName=cmtracker_nsedef.css")
#     driver2.find_element_by_xpath('//*[@id="tab52"]/a').click()
#     WebDriverWait(driver2, 60).until(EC.presence_of_element_located((By.XPATH, "/html/body/table/tbody/tr[4]/td/table/tbody/tr/td[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td[2]/form/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/nobr/b")))
#     ele = driver2.find_element_by_xpath('/html/body/table/tbody/tr[4]/td/table/tbody/tr/td[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td[2]/form/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr/td[4]/nobr/input')
#     ele.clear()
#     ele.send_keys(end_day)
#     ele = driver2.find_element_by_xpath('/html/body/table/tbody/tr[4]/td/table/tbody/tr/td[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td[2]/form/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr/td[8]/nobr/input')
#     ele.clear()
#     ele.send_keys(start_day)
#     driver2.find_element_by_xpath('/html/body/table/tbody/tr[4]/td/table/tbody/tr/td[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td[2]/form/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr/td[9]/input').click()
#     WebDriverWait(driver2, 60).until(EC.presence_of_element_located((By.XPATH,'/html/body/table/tbody/tr[4]/td/table/tbody/tr/td[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td[2]/form/table/tbody/tr/td[1]/table/tbody/tr[4]/td/a')))
#     tuname = driver2.find_element_by_xpath('/html/body/table/tbody/tr[4]/td/table/tbody/tr/td[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td[2]/form/table/tbody/tr/td[1]/table/tbody/tr[4]/td/a').get_attribute("href")
#     tuname = tuname.split(r"/")[-1]
#     if (not os.path.exists(os.getcwd()+"\\temp\\"+tuname)):
#         driver2.find_element_by_xpath('/html/body/table/tbody/tr[4]/td/table/tbody/tr/td[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td[2]/form/table/tbody/tr/td[1]/table/tbody/tr[4]/td/a').click()
    
# #       shutil.move(os.getcwd()+"\\"+tuname,os.getcwd()+"\\temp\\"+tuname) 
#     try:
#         while (not os.path.exists(os.getcwd()+"\\temp\\"+tuname)):
#             time.sleep(0.1)
#         gframe = pandas.read_csv(os.getcwd()+"\\temp\\"+tuname)
#     except:
#         time.sleep(0.2)
#         gframe = pandas.read_csv(os.getcwd()+"\\temp\\"+tuname)
#     tuurn_avg =[]
#     pframe= gframe.iloc[0:,[0,1]]
#     tuurn = list(pframe.iloc[:,1])
#     for i in range(pframe.iloc[:,0].size):
#         if ( i<4):
#             tuurn_avg.append(0)
#         else:
#             tuurn_avg.append((pframe.iloc[i-4][1]+pframe.iloc[i][1]+pframe.iloc[i-1][1]+pframe.iloc[i-2][1]+pframe.iloc[i-3][1])/5 )

# #     print(daylist,"dates /for turnover of each company data",gframe.iloc[0:,0])
    
    totalRecord = 0
#     i=1
    dataForCSV=[]
    
    
    elem = driver.find_element_by_name("symbol")
    elem.clear()
    elem.send_keys(company)
    driver.find_element_by_xpath("//select[@name='dataType']/option[text()='Security-wise Price volume & Deliverable position data']").click()
    driver.find_element_by_xpath("//select[@name='series']/option[text()='EQ']").click() 
    TE = "3 months"
    driver.find_element_by_xpath("//select[@name='dateRange']/option[text()='"+TE+"']").click() 
    driver.find_element_by_xpath('//*[@id="get"]').click()
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[3]/div[2]/div[1]/div[3]/div/div[3]/table/tbody")))
    while driver.find_element_by_xpath("/html/body/div[2]/div[3]/div[2]/div[1]/div[3]/div/div[3]/table/tbody/tr[2]/td[1]").text != company:
            print(company,driver.find_element_by_xpath('/html/body/div[2]/div[3]/div[2]/div[1]/div[3]/div/div[3]/table/tbody/tr[2]/td[1]').text)
            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[3]/div[2]/div[1]/div[3]/div/div[3]/table/tbody")))
    

    with open('temp.csv', 'w', newline='') as csvfile:
        spamwriter = csv.writer(csvfile,quotechar='|', quoting=csv.QUOTE_MINIMAL)
        for a in (driver.find_element_by_xpath('//*[@id="csvContentDiv"]').get_attribute("innerHTML")).replace('"',"").replace(" ","").split(":"):
            ne = list(a.split(","))
            spamwriter.writerow(ne)
        
    a = pandas.read_csv("temp.csv")
    a.dropna(axis=0,inplace=True)
    i = -1
    while(i>-1*(rcd_update+5)):
#         print(i)
        l = []
        l.append(a.iloc[i].Date)
        l.append(a.iloc[i].OpenPrice) 
        l.append(a.iloc[i].HighPrice)
        l.append(a.iloc[i].LowPrice)
        l.append(a.iloc[i].ClosePrice)
#       change in open and close price
        l.append(((a.iloc[i].ClosePrice - a.iloc[i-1].ClosePrice)/a.iloc[i-1].ClosePrice*100).round(2))
        l.append(a.iloc[i].AveragePrice)
        l.append(((a.iloc[i].AveragePrice+a.iloc[i-1].AveragePrice+a.iloc[i-2].AveragePrice+a.iloc[i-3].AveragePrice+a.iloc[i-4].AveragePrice)/5).round(2))
        l.append(a.iloc[i].Turnover)
        l.append(((a.iloc[i].Turnover+a.iloc[i-1].Turnover+a.iloc[i-2].Turnover+a.iloc[i-3].Turnover+a.iloc[i-4].Turnover)/5).round(2))
        l.append(a.iloc[i].TotalTradedQuantity)
        l.append(a.iloc[i].TotalTradedQuantity-a.iloc[i-1].TotalTradedQuantity)
        l.append(((a.iloc[i].TotalTradedQuantity-a.iloc[i-1].TotalTradedQuantity)/a.iloc[i-1].TotalTradedQuantity*100).round(2))
        l.append(((a.iloc[i].TotalTradedQuantity+a.iloc[i-1].TotalTradedQuantity+a.iloc[i-2].TotalTradedQuantity+a.iloc[i-3].TotalTradedQuantity+a.iloc[i-4].TotalTradedQuantity)/5).round(2))
        dataForCSV.append(l) 
        i-=1
#     since df are in order lastes -> oldest
#     and dataForCSV in oldest -> latest
    dataForCSV = dataForCSV[::-1]
    Future_Data = []
    Option_Data = []
    for i,df in enumerate(all_days_df[::-1]):
        
        
        df = df[df.SYMBOL.isin([company])]
        df2 = df[df.INSTRUMENT.isin(["FUTSTK"])]
        sum1 = df2["OPEN_INT"].sum()
        sum2 = df2["CONTRACTS"].sum()
        sum5 = df2['VAL_INLAKH'].sum()
        df3 = df[df.INSTRUMENT.isin(["OPTSTK"])]
        sum3 = df3["OPEN_INT"].sum()
        sum4 = df3["CONTRACTS"].sum()
        sum6 = df3["VAL_INLAKH"].sum()
        del(df)
        del(df2)
        del(df3)
        if(i<4):
            dataForCSV[i]= dataForCSV[i][:-4]+dataForCSV[i][-4:]+["",sum1,0,0,0,sum5,0,sum2,0,0,0,"",sum3,0,0,0,sum6,0,sum4,0,0,0]
            Future_Data.append([sum1,0,0,0,sum5,0,sum2,0,0,0])
            Option_Data.append([sum3,0,0,0,sum6,0,sum4,0,0,0])
        else:
#             print(i)
            OIchange = (sum1 - float(Future_Data[i-1][0]))
            OIchange_ = round(OIchange*100/float(Future_Data[i-1][0]),2)
            OIAvg =int((sum1+float(Future_Data[i-1][0])+float(Future_Data[i-2][0])+float(Future_Data[i-3][0])+float(Future_Data[i-1][0]))/5)
            cont_change = (sum2 - float(Future_Data[i-1][6]))
            cont_change_ = round(cont_change*100/float(Future_Data[i-1][6]),2)
            cont_Avg =int((sum2+float(Future_Data[i-1][6])+float(Future_Data[i-2][6])+float(Future_Data[i-3][6])+float(Future_Data[i-4][6]))/5)

            FutTur_Avg =  int((sum5+float(Future_Data[i-1][4])+float(Future_Data[i-2][4])+float(Future_Data[i-3][4])+float(Future_Data[i-4][4]))/5)
            
            oichange = (sum3 -float(Option_Data[i-1][0]))
            oichange_ = round(oichange*100/float(Option_Data[i-1][0]),2)
            oiAvg =int((sum3+float(Option_Data[i-1][0])+float(Option_Data[i-2][0])+float(Option_Data[i-3][0])+float(Option_Data[i-1][0]))/5)
            con_change = (sum4 - float(Option_Data[i-1][6]))
            con_change_ = round(con_change*100/float(Option_Data[i-1][6]),2)
            con_Avg =int((sum4+float(Option_Data[i-1][6])+float(Option_Data[i-2][6])+float(Option_Data[i-3][6])+float(Option_Data[i-4][6]))/5)
            OptTur_Avg = int((sum6+float(Option_Data[i-1][4])+float(Option_Data[i-2][4])+float(Option_Data[i-3][4])+float(Option_Data[i-4][4]))/5)
            Future_Data.append([sum1,OIchange,OIchange_,OIAvg,sum5,FutTur_Avg,sum2,cont_change,cont_change_,cont_Avg])
            Option_Data.append([sum3,oichange,oichange_,oiAvg,sum6,OptTur_Avg,sum4,con_change,con_change_,con_Avg])
            
            dataForCSV[i]= dataForCSV[i][:-4]+dataForCSV[i][-4:]+["",sum1,OIchange,OIchange_,OIAvg,sum5,FutTur_Avg,sum2,cont_change,cont_change_,cont_Avg,"",sum3,oichange,oichange_,oiAvg,sum6,OptTur_Avg,sum4,con_change,con_change_,con_Avg]
            
    if(create_new): 
        workbook = xlsxwriter.Workbook(os.getcwd() + "\\company\\"+company+'.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.set_row(2, 60)
        merge_format1 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#59db99'})
        merge_format2 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#d0aa5b'})
        merge_format3 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'border_color':'black',
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '1fa1b8'})

        merge_format4 = workbook.add_format({
            'font_name':"Calibri",
            'border': 1,
            'align': 'center',
            "font_size" : 48,
            'valign': 'vcenter',
            'bg_color': '#F2F2F2'})

        
        worksheet.merge_range('A1:AJ1', company, merge_format4)
        # worksheet.merge_range('B3:L3',"")
        worksheet.merge_range('B3:N3', 'Cash', merge_format1)
        worksheet.merge_range('P3:Y3', 'Future', merge_format2)
        worksheet.merge_range('AA3:AJ3', 'Option',merge_format3)


        row = 4
        for i in range(len(feildName)):
                worksheet.write(3,i,feildName[i])

        for element in dataForCSV[4:]:
            for i in range(len(element)):
                worksheet.write(row,i,element[i])
            row+=1
        workbook.close()
    elif(continue_from !=0) :
        p = pandas.read_excel(os.getcwd()+"\\company\\"+company+".xlsx").drop([0,1])
        p.columns = p.iloc[0]
        p = p.drop([2,])
        Add_row = []
        for i in range(p.iloc[:continue_from-3,0].size):
            row_data = p.iloc[i,:]
        #     Add_row.append(row_data)
            new_row = []
            for element in row_data:
                if(str(element) == 'nan'):
                    new_row.append("")
                else:
                    new_row.append(element)
            Add_row.append(new_row)
        final_row_data = Add_row+dataForCSV[4:]
        workbook = xlsxwriter.Workbook(os.getcwd() + "\\company\\"+company+'.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.set_row(2, 60)
        merge_format1 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#59db99'})
        merge_format2 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#d0aa5b'})
        merge_format3 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'border_color':'black',
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '1fa1b8'})

        merge_format4 = workbook.add_format({
            'font_name':"Calibri",
            'border': 1,
            'align': 'center',
            "font_size" : 48,
            'valign': 'vcenter',
            'bg_color': '#F2F2F2'})


        worksheet.merge_range('A1:AJ1', company, merge_format4)
        # worksheet.merge_range('B3:L3',"")
        worksheet.merge_range('B3:N3', 'Cash', merge_format1)
        worksheet.merge_range('P3:Y3', 'Future', merge_format2)
        worksheet.merge_range('AA3:AJ3', 'Option',merge_format3)

        row = 4
        for i in range(len(feildName)):
                worksheet.write(3,i,feildName[i])

        for element in final_row_data:
            for i in range(len(element)):
                worksheet.write(row,i,element[i])
            row+=1
        workbook.close()
        
    else:
        wb = load_workbook((os.getcwd() + "\\company\\"+company+'.xlsx'))
        ws = wb.worksheets[0]

        for row_data in dataForCSV[4:]:
            ws.append(row_data)
        wb.save((os.getcwd() + "\\company\\"+company+'.xlsx'))
    print(company," Data computed")
# driver2.close()        

# from selenium import webdriver
# driver = webdriver.Firefox(executable_path =local_dir,firefox_profile=fp)
# code is for Turnover Xlsx
col= ["Date","1","IndexFuture","3","4","5","StockFuture","7","8","IndexOption","10","11","12","StockOption","14","15","16","17"]
finalColoumn =["DATE","INDEX FUTURES(Cr)","% CHG","5 DAYS AVG","","INDEX OPTIONS(Cr)","% CHG","5 DAYS AVG","PCR","","STOCK FUTURES(Cr)","% CHG","5 DAYS AVG","","STOCK OPTIONS(Cr)","% CHG","5 DAYS AVG","PCR"]
month = ""
finalData = []
month_list = []
f_name = []
for element in daylist:
#     print(daylist)
#     print(element[3:6])
    if element[3:6] not in month_list:
        month_list.append(element[3:6])
        month = element[3:6]
        index = months.index(month.upper())
        f_name.append("fo_turn"+element[3:6].lower()+element[7:]+".htm")
   
        
for l,nm in enumerate(f_name):
        op = l+2
        driver.get('https://www1.nseindia.com//products/content/derivatives/equities/'+nm)
        data = driver.find_element_by_xpath('/html/body/table').text[394:].split("\n")
        for i,each in enumerate(data):
            data[i] = each.split(" ")
        with open('temp'+str(op)+'.csv', 'w', newline='') as csvfile:
            spamwriter = csv.writer(csvfile)
            spamwriter.writerow(col)
            for a in data:
                spamwriter.writerow(a)
    
df = pandas.read_csv("temp2.csv")
if(len(f_name)==2):
    df = pandas.concat([pandas.read_csv("temp3.csv"),df])
elif(len(f_name)==3):
    df = pandas.concat([pandas.read_csv("temp3.csv"),df])
    df = pandas.concat([pandas.read_csv("temp4.csv"),df])
i=-1

# print(df)
while df.iloc[i].Date.upper() in daylist:
#     print(i)
    l = []
    l.append(df.iloc[i].Date.upper())
    l.append(df.iloc[i].IndexFuture)
    if(i>-1*rcd_update-1):
        l.append(((df.iloc[i].IndexFuture-df.iloc[i-1].IndexFuture)/df.iloc[i-1].IndexFuture*100).round(2))
        l.append(((df.iloc[i].IndexFuture+df.iloc[i-1].IndexFuture+df.iloc[i-2].IndexFuture+df.iloc[i-3].IndexFuture+df.iloc[i-4].IndexFuture)/5).round(2))
    else:
        l.append(0)
        l.append(0)
    l.append("")
    l.append(df.iloc[i].IndexOption)
    if(i>-1*rcd_update-1):
        l.append(((df.iloc[i].IndexOption-df.iloc[i-1].IndexOption)/df.iloc[i-1].IndexOption*100).round(2))
        l.append(((df.iloc[i].IndexOption+df.iloc[i-1].IndexOption+df.iloc[i-2].IndexOption+df.iloc[i-3].IndexOption+df.iloc[i-4].IndexOption)/5).round(2))
    else:
        l.append(0)
        l.append(0)
    l.append(df.iloc[i]["10"])
    l.append("")
    l.append(df.iloc[i].StockFuture)
    if(i>-1*rcd_update-1):
        l.append(((df.iloc[i].StockFuture-df.iloc[i-1].StockFuture)/df.iloc[i-1].StockFuture*100).round(2))
        l.append(((df.iloc[i].StockFuture+df.iloc[i-1].StockFuture+df.iloc[i-2].StockFuture+df.iloc[i-3].StockFuture+df.iloc[i-4].StockFuture)/5).round(2))
    else:
        l.append(0)
        l.append(0)
    l.append("")
    l.append(df.iloc[i].StockOption)
    if(i>-1*rcd_update-1):
        l.append(((df.iloc[i].StockOption-df.iloc[i-1].StockOption)/df.iloc[i-1].StockOption*100).round(2))
        l.append(((df.iloc[i].StockOption+df.iloc[i-1].StockOption+df.iloc[i-2].StockOption+df.iloc[i-3].StockOption+df.iloc[i-4].StockOption)/5).round(2))
    else:
        l.append(0)
        l.append(0)
    l.append(df.iloc[i]["14"])
    i-=1
    finalData.append(l)
finalData= finalData[::-1]

# prin
create_new =1
continue_from = 0

if (os.path.exists(os.getcwd()+"\\turnover\\"+"Turnover"+".xlsx")):
        edit_date = daylist[rcd_update-1]
        append_ = 1
        create_new = 0
        ty = pandas.read_excel(os.getcwd()+"\\turnover\\"+"Turnover"+".xlsx").iloc[:,0]
        for i in range(ty.size):
                print(ty.iloc[i])
                if(i>2 and str(ty.iloc[i]).upper() == edit_date):
                    append_ = 0
                    continue_from = i
                    break


# print(edit_date)
Add_row = []
if(continue_from>0):
    p = pandas.read_excel((os.getcwd()+"\\turnover\\"+"Turnover"+".xlsx"))

    #     print("\n\n data",p.iloc[3:continue_from+1,:])
    for i in range(continue_from - 3):
#         row_data = p.iloc[3+i,:]
#         print(row_data)
        new_row = []
        row_data = list(p.iloc[3+i,:])
        for element in row_data:
            if(str(element) == 'nan'):
                new_row.append("")
            else:
                new_row.append(element)
        Add_row.append(new_row[:-1])
#     print(Add_row)
if(create_new):    
    workbook = xlsxwriter.Workbook(os.getcwd()+"//turnover"+"//Turnover"+'.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.set_row(2, 60)
    merge_format1 = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#59db99'})
    merge_format2 = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#d0aa5b'})
    merge_format3 = workbook.add_format({
        'bold': 1,
        'border': 1,
        'border_color':'black',
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '1fa1b8'})

    merge_format4 = workbook.add_format({
        'font_name':"Calibri",
        'border': 1,
        'align': 'center',
        "font_size" : 48,
        'valign': 'vcenter',
        'bg_color': '#F2F2F2'})


    worksheet.merge_range('A1:R1',"Turnover", merge_format4)
    worksheet.merge_range('B3:D3', 'Index Future', merge_format1)
    worksheet.merge_range('F3:I3', 'Index Option', merge_format2)
    worksheet.merge_range('K3:M3', 'Stock Future',merge_format1)
    worksheet.merge_range('O3:R3', 'Stock Option',merge_format2)
    row = 4
    for i in range(len(finalColoumn)):
            worksheet.write(3,i,finalColoumn[i])

    for element in finalData[-10:]:
        for i in range(len(finalColoumn)):
            if(i!=0 and i!=4 and i!=9 and i!=13  ):
                try:
                    worksheet.write(row,i,float(element[i]))
                except:
                    print(element[i],"error",row,i)
                    break
            else:
                worksheet.write(row,i,element[i])
        row+=1
    workbook.close()
    print("Turnover Updated")
elif(append_!=1):
    workbook = xlsxwriter.Workbook(os.getcwd()+"//turnover"+"//Turnover"+'.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.set_row(2, 60)
    merge_format1 = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#59db99'})
    merge_format2 = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#d0aa5b'})
    merge_format3 = workbook.add_format({
        'bold': 1,
        'border': 1,
        'border_color':'black',
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '1fa1b8'})

    merge_format4 = workbook.add_format({
        'font_name':"Calibri",
        'border': 1,
        'align': 'center',
        "font_size" : 48,
        'valign': 'vcenter',
        'bg_color': '#F2F2F2'})


    worksheet.merge_range('A1:R1',"Turnover", merge_format4)
    # worksheet.merge_range('B3:L3',"")
    worksheet.merge_range('B3:D3', 'Index Future', merge_format1)
    worksheet.merge_range('F3:I3', 'Index Option', merge_format2)
    worksheet.merge_range('K3:M3', 'Stock Future',merge_format1)
    worksheet.merge_range('O3:R3', 'Stock Option',merge_format2)
    row = 4
    for i in range(len(finalColoumn)):
            worksheet.write(3,i,finalColoumn[i])
    
    for element in Add_row:
        for i in range(len(element)):
            if(i!=0 and i!=4 and i!=9 and i!=13  ):
                try:
                    worksheet.write(row,i,float(element[i]))
                except:
                    print(element[i],"error",row,i)
                    break
            else:
                worksheet.write(row,i,element[i])
        row+=1
#     row+= 5
#     print(finalData)
#     print(finalData[-1*rcd_update])
    for element in finalData[-1*rcd_update:]:
        for i in range(len(finalColoumn)):
            if(i!=0 and i!=4 and i!=9 and i!=13  ):
                try:
                    worksheet.write(row,i,float(element[i]))
                except:
                    print(element[i],"error",row,i)
                    break
            else:
                worksheet.write(row,i,element[i])
        row+=1
    workbook.close()
    
else:
    wb = load_workbook((os.getcwd() + '\\turnover\\Turnover.xlsx'))
    ws = wb.worksheets[0]

    for row_data in finalData[4:]:
        ws.append(row_data)
    wb.save((os.getcwd() + '\\turnover\\Turnover.xlsx'))


driver.get('https://www1.nseindia.com/products/content/equities/indices/historical_index_data.htm')
driver.find_element_by_xpath('//*[@id="toDate"]').send_keys(start_day)
driver.find_element_by_xpath('//*[@id="fromDate"]').send_keys(end_day)
# 

for o,company_ in enumerate(["NIFTY","BANKNIFTY"]):
#     print(df.iloc[:,-2])
    driver.find_element_by_xpath('/html/body/div[2]/div[3]/div[2]/div[1]/div[4]/div/div[1]/div/div[2]/select/optgroup['+str(o+1)+']/option['+str(o+1)+']').click()
    driver.find_element_by_xpath('/html/body/div[2]/div[3]/div[2]/div[1]/div[4]/div/div[1]/div/div[2]/select/optgroup[1]').click()
    driver.execute_script(' loadindicesdata()')
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div[3]/div[2]/div[1]/div[4]/div/div[2]/table")))
    lol = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div[2]/div[1]/div[4]/div/div[2]/table').get_attribute('innerHTML')



    dataForCSV = []
    soup = BeautifulSoup(lol)
    table = soup.select_one("tbody")
    headers = [th.text for th in table.select("tr th")]
   
    ya = [[td.text for td in row.find_all("td")] for row in table.select("tr + tr")][2:-1]
#     print("printing the fetched table\n",ya)
    for p,data_ in enumerate(ya):
        if data_[5]=='-' or data_[5]=='_':
            data_[5]= "0"
            data_[6]= "0"
            
        temp_data = []
        for i,j in enumerate(data_[:5]+[data_[-1]]):
            if(i>0):
                temp_data.append(float(j.replace(" ","")))
            else:
                temp_data.append(j.replace(" ",""))
        tm = temp_data.pop()
        if(len(dataForCSV)>1):
            temp_data.append(round((float(temp_data[4]) -float(dataForCSV[p-1][4]))/float(dataForCSV[p-1][4])*100,2))
        else:
            temp_data.append(0)

        temp_data.append(tm)
        temp_data.append(float(data_[5]))

        dataForCSV.append(temp_data)
    dataForCSV = dataForCSV[-1*rcd_update-4:]
#     print(len(dataForCSV),dataForCSV[0] )
    
    for i,df in  enumerate(all_days_df[::-1]):
#         i=j
#         print(df.iloc[5,-2],dataForCSV[i][0])
        df = df[df.SYMBOL.isin([company_])]
        df1 = df[df.INSTRUMENT.isin(['FUTIDX'])]
        df2 = df[df.INSTRUMENT.isin(['OPTIDX'])]
        
        sum2 = df1.CONTRACTS.sum()
        sum1 = df1.OPEN_INT.sum()
        sum4 = df2.CONTRACTS.sum()
        sum3 = df2.OPEN_INT.sum()
        
        if i<4:
            vol_df = 0
            vol_dif_per = 0
            vol_avg = 0
            turn_dif = 0
            turn_avg = 0
        else:
            day_skip = 0
            for j in range(3):
                    if(float(dataForCSV[i-2-j][6])==0.0):
                        day_skip+=1
                        
            if(float(dataForCSV[i-1][6])==0.0):
                turn_dif = 0
                day_skip = 1
                for j in range(3):
                    if(float(dataForCSV[i-2-j][6])==0.0):
                        day_skip+=1
                turn_avg = round((float(dataForCSV[i][6])+float(dataForCSV[i-1][6])+float(dataForCSV[i-2][6])+float(dataForCSV[i-3][6])+float(dataForCSV[i-4][6]))/(5-day_skip),2)        
                vol_df = 0
                vol_dif_per = 0
                print(day_skip,i)
                vol_avg = round((float(dataForCSV[i][7])+float(dataForCSV[i-1][9])+float(dataForCSV[i-2][9])+float(dataForCSV[i-3][9])+float(dataForCSV[i-4][9]))/(5-day_skip),2)
            elif(float(dataForCSV[i][6])==0.0):
                turn_dif = 0
                day_skip = 1
                for j in range(4):
                    if(float(dataForCSV[i-1-j][6])==0.0):
                        day_skip+=1
                turn_avg = round((float(dataForCSV[i][6])+float(dataForCSV[i-1][6])+float(dataForCSV[i-2][6])+float(dataForCSV[i-3][6])+float(dataForCSV[i-4][6]))/(5-day_skip),2)        
                vol_df = 0
                vol_dif_per = 0
                print(day_skip,i)
                vol_avg = round((float(dataForCSV[i][7])+float(dataForCSV[i-1][9])+float(dataForCSV[i-2][9])+float(dataForCSV[i-3][9])+float(dataForCSV[i-4][9]))/(5-day_skip),2)
        
            else:
                turn_dif = round((float(dataForCSV[i][6])-float(dataForCSV[i-1][6]))/float(dataForCSV[i-1][6])*100,2)
                turn_avg = round((float(dataForCSV[i][6])+float(dataForCSV[i-1][6])+float(dataForCSV[i-2][6])+float(dataForCSV[i-3][6])+float(dataForCSV[i-4][6]))/(5-day_skip),2)      
                vol_df = (float(dataForCSV[i][7])-float(dataForCSV[i-1][9]))
                vol_dif_per = round((float(dataForCSV[i][7])-float(dataForCSV[i-1][9]))/float(dataForCSV[i-1][9])*100,2)
                vol_avg = round((float(dataForCSV[i][7])+float(dataForCSV[i-1][9])+float(dataForCSV[i-2][9])+float(dataForCSV[i-3][9])+float(dataForCSV[i-4][9]))/(5-day_skip),2)
#         print(dataForCSV[i])
        if i<4 : 
            dataForCSV[i]= dataForCSV[i][:-1]+[turn_dif,turn_avg]+[dataForCSV[i][-1]]+[vol_df,vol_dif_per,vol_avg]+["",sum1,0,0,0,sum2,0,0,0,"",sum3,0,0,0,sum4,0,0,0]

        else:
            OIchange = (sum1 - float(dataForCSV[i-1][14]))
            OIchange_ = round(OIchange*100/float(dataForCSV[i-1][14]),2)
            OIAvg =int((sum1+float(dataForCSV[i-1][14])+float(dataForCSV[i-2][14])+float(dataForCSV[i-3][14])+float(dataForCSV[i-4][14]))/5)
            cont_change = (sum2 - float(dataForCSV[i-1][18]))
            cont_change_ = round(cont_change*100/float(dataForCSV[i-1][18]),2)
            cont_Avg =int((sum2+float(dataForCSV[i-1][18])+float(dataForCSV[i-2][18])+float(dataForCSV[i-3][18])+float(dataForCSV[i-4][18]))/5)


            oichange = (sum3 - float(dataForCSV[i-1][23]))
            oichange_ = round(oichange*100/float(dataForCSV[i-1][23]),2)
            oiAvg =int((sum3+float(dataForCSV[i-1][23])+float(dataForCSV[i-2][23])+float(dataForCSV[i-3][23])+float(dataForCSV[i-4][23]))/5)
            con_change = (sum4 - float(dataForCSV[i-1][27]))
            con_change_ = round(con_change*100/float(dataForCSV[i-1][27]),2)
            con_Avg =int((sum4+float(dataForCSV[i-1][27])+float(dataForCSV[i-2][27])+float(dataForCSV[i-3][27])+float(dataForCSV[i-4][27]))/5)
            dataForCSV[i]= dataForCSV[i][:-1]+[turn_dif,turn_avg]+[dataForCSV[i][-1]]+[vol_df,vol_dif_per,vol_avg]+["",sum1,OIchange,OIchange_,OIAvg,sum2,cont_change,cont_change_,cont_Avg,"",sum3,oichange,oichange_,oiAvg,sum4,con_change,con_change_,con_Avg]
#             print("appended",dataForCSV[i][0])
#     create_new =1
#     continue_from = 0
#     if (os.path.exists(os.getcwd()+ "/nifty/"+company_+'.xlsx')):
#         last_saved_date = pandas.read_excel(os.getcwd()+ "/nifty/"+company_+'.xlsx').iloc[:,0].iloc[-1]
#         data_should_append =1
#         create_new = 0
#         continue_from = 5
#         for i in range(len(dataForCSV[5:])):
#                 if(dataForCSV[5+i][0] == last_saved_date):
#                     create_new = 0
#                     continue_from = 6+i
#                     break
                    
    create_new =1
    continue_from = 0

    if (os.path.exists(os.getcwd()+ "\\nifty\\"+company_+'.xlsx')):
        edit_date = daylist[rcd_update-1]
        append_ = 1
        create_new = 0
        ty = pandas.read_excel(os.getcwd()+ "\\nifty\\"+company_+'.xlsx').iloc[:,0]
        for i in range(ty.size):
#                 print(ty.iloc[i])
#                 print(str(ty.iloc[i]).upper())
                if(i>2 and str(ty.iloc[i]).upper() == edit_date):
#                     print(str(ty.iloc[i]).upper())
                    append_ = 0
                    continue_from = i
                    break
#     print(continue_from,ty,edit_date)
    Add_row = []
    if(continue_from>0):
        p = pandas.read_excel(os.getcwd()+ "\\nifty\\"+company_+'.xlsx')
        for i in range(continue_from - 3):
            new_row = []
            row_data = list(p.iloc[3+i,:])
#             print(row_data)
            for element in row_data:
                if(str(element) == 'nan'):
                    new_row.append("")
                else:
                    new_row.append(element)
            Add_row.append(new_row[:])
# #     print(Add_row)

# # print(dataForCSV)
    if(create_new):            
        workbook = xlsxwriter.Workbook(os.getcwd()+ "\\nifty\\"+company_+'.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.set_row(2, 60)
        merge_format1 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#59db99'})
        merge_format2 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#d0aa5b'})
        merge_format3 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'border_color':'black',
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '1fa1b8'})

        merge_format4 = workbook.add_format({
            'font_name':"Calibri",
            'border': 1,
            'align': 'center',
            "font_size" : 48,
            'valign': 'vcenter',
            'bg_color': '#F2F2F2'})


        worksheet.merge_range('A1:AE1', company_, merge_format4)
        worksheet.merge_range('B3:M3', 'Cash', merge_format1)
        worksheet.merge_range('O3:V3', 'Future', merge_format2)
        worksheet.merge_range('X3:AE3', 'Option',merge_format3)

        row = 4
        for i in range(len(nifyty_header)):
                worksheet.write(3,i,nifyty_header[i])

        for element in dataForCSV[4:]:

            for i in range(len(nifyty_header)):
                if(i!=0 and i!=13 and i!= 22):
                    try:
                        worksheet.write(row,i,float(element[i]))
                    except:
                        print(element[i],"error",row,i)
                        break
                else:
                    worksheet.write(row,i,element[i])
            row+=1
        workbook.close()
    elif(append_ == 0):
        workbook = xlsxwriter.Workbook(os.getcwd()+ "\\nifty\\"+company_+'.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.set_row(2, 60)
        merge_format1 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#59db99'})
        merge_format2 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '#d0aa5b'})
        merge_format3 = workbook.add_format({
            'bold': 1,
            'border': 1,
            'border_color':'black',
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': '1fa1b8'})

        merge_format4 = workbook.add_format({
            'font_name':"Calibri",
            'border': 1,
            'align': 'center',
            "font_size" : 48,
            'valign': 'vcenter',
            'bg_color': '#F2F2F2'})

        
        worksheet.merge_range('A1:AE1', company_, merge_format4)
        worksheet.merge_range('B3:M3', 'Cash', merge_format1)
        worksheet.merge_range('O3:V3', 'Future', merge_format2)
        worksheet.merge_range('X3:AE3', 'Option',merge_format3)

        
        row = 4
        for i in range(len(nifyty_header)):
                worksheet.write(3,i,nifyty_header[i])

        for element in Add_row:
            for i in range(len(element)):
                if(i!=0 and i!=13 and i!= 22):
                    try:
                        worksheet.write(row,i,float(element[i]))
                    except:
                        print(element[i],"error",row,i)
                        break
                else:
                    worksheet.write(row,i,element[i])
            row+=1
            
            
        for element in dataForCSV[4:]:

            for i in range(len(nifyty_header)):
                if(i!=0 and i!=13 and i!= 22):
                    try:
                        worksheet.write(row,i,float(element[i]))
                    except:
                        print(i,"gives error")
#                         print(element[i],"error",row,i)
                        break
                else:
                    worksheet.write(row,i,element[i])
            row+=1
                
        workbook.close()



    
    else:
        print("data Appended")
        wb = load_workbook(os.getcwd()+ "\\nifty\\"+company_+'.xlsx')
        ws = wb.worksheets[0]

        for row_data in dataForCSV[4:]:
            ws.append(row_data)
        wb.save(os.getcwd()+ "\\nifty\\"+company_+'.xlsx')
print("NIFTY AND BANKNIFTY Updated")

if os.path.exists(os.getcwd()+"//temp1.csv"):
    os.remove(os.getcwd()+"//temp1.csv")
if os.path.exists(os.getcwd()+"//temp2.csv"):
    os.remove(os.getcwd()+"//temp2.csv")
if os.path.exists(os.getcwd()+"//temp3.csv"):
    os.remove(os.getcwd()+"//temp3.csv")
if os.path.exists(os.getcwd()+"//temp.csv"):
    os.remove(os.getcwd()+"//temp.csv")

driver.close()
